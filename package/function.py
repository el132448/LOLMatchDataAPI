import requests
import pandas as pd
import datetime
import time
import os.path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
# from UliPlot.XLSX import auto_adjust_xlsx_column_width

class Setting:
    setting = ["大涼粉",'tw2','SEA',100,450,'RGAPI-d0df3f5e-06ee-4b76-ac3c-a2c93d3651aa',
                "gfh2NBRvQfamzlVNBVKT7V7AV_4N5aBiPWQODvOvAkkNF6zVbr1D0cmKlWd2uOr5sNGYYGdu80G8-Q",
                1623801600,int(time.time())]

class Data: 
    def __init__(self, summoner_name: str, region: str, mass_region, no_games, queue_id, api_key, puuid, startTime, endTime):
        self.summoner_name = summoner_name
        self.region = region
        self.mass_region = mass_region
        self.no_games = no_games
        self.queue_id = queue_id
        self.api_key = api_key
        self.puuid = puuid
        self.startTime = startTime
        self.endTime = endTime
        self.df = pd.DataFrame()
        self.stat = {}
    #=============================================================================================================
    #=============================================================================================================
    def checkApiKey(self):
        i = '100' # 100 means keep on looping
        while i != 0: # loop until API key is vaild
            api_url = ("https://" + self.region + ".api.riotgames.com/lol/platform/v3/champion-rotations?api_key=" + self.api_key)
            resp = requests.get(api_url)
            # If API key expired
            print(resp)
            if str(resp) != '<Response [200]>':
                print(str(resp))
                self.api_key = input('Please input a vaild API key. (enter 0 to exit)\n')
                if self.api_key == '0':
                    exit()
            elif str(resp) == '<Response [200]>':
                i = 0 # end the loop
                print('api key is vaild!')
    #=============================================================================================================
    #=============================================================================================================
    # gets the puuid, given a summoner name and region
    def get_puuid(self):

        api_url = ("https://" + self.region + ".api.riotgames.com/lol/summoner/v4/summoners/by-name/" + self.summoner_name + "?api_key=" + self.api_key)
        resp = requests.get(api_url)

        player_info = resp.json()
        self.puuid = player_info['puuid']
        print('puuid get!')
    #=============================================================================================================
    #=============================================================================================================
    def updateSummonerName(self):
        api_url = ("https://" + self.region + ".api.riotgames.com/lol/summoner/v4/summoners/by-puuid/"+ self.puuid + "?api_key=" + self.api_key)
        resp = requests.get(api_url)
        print(api_url)
        player_info = resp.json()
        print(player_info)
        self.summoner_name = player_info['name']
        print('summoner_name get!')
    #=============================================================================================================
    #=============================================================================================================
    # get a list of all the match IDs
    def get_match_ids(self, endTime):
        api_url = (
            "https://" +
            self.mass_region +
            ".api.riotgames.com/lol/match/v5/matches/by-puuid/" +
            self.puuid + 
            "/ids?startTime="+ str(self.startTime) + "&endTime="+ str(endTime) +
            "&queue=" + str(self.queue_id) +
            "&start=0&count=" + str(self.no_games) + 
            "&api_key=" + self.api_key
        )
        resp = requests.get(api_url)
        match_ids = resp.json()
        print(str(resp))
        if str(resp) == '<Response [200]>':
            print('match_ids get!')
            return match_ids
        else:
            return []
    #=============================================================================================================
    #=============================================================================================================
    # From a given match ID, get the data about the game
    def get_match_data(self, match_id):
        api_url = (
            "https://" + 
            self.mass_region + 
            ".api.riotgames.com/lol/match/v5/matches/" +
            match_id + 
            "?api_key=" + 
            self.api_key
        )
        
        # need "while" so that continuously loop until successful
        while True:
            resp = requests.get(api_url)
            
            # whenever we see a 429, sleep for 10 seconds and then restart from the top of the "while" loop
            if resp.status_code == 429:
                print("Rate Limit hit, sleeping for 10 seconds (100 requests every 2 minutes)")
                time.sleep(10)
                # continue means start the loop again
                continue
                
            # if resp.status_code isn't 429, then we carry on to the end of the function and return the data
            match_data = resp.json()
            return match_data   
    #=============================================================================================================
    #=============================================================================================================
    def gather_all_Data(self, match_ids, i):
        # initialise an empty dictionary to store data for each game
        data = {
            'matchId': [],
            'gameStartTime': [],
            'date': [],
            'time': [],
            'gameLength': [],
            'mode': [],
            'gameVersion': [],
            'team' : [],
            'win': [],
            'earlySurrender': [],
            'surrender': [],
            'summonerName': [],
            'summonerLevel': [],
            'champion': [],
            'kills': [],
            'deaths': [],
            'assists': [],
            'KDA': [],
            'quadraKills': [],
            'pentaKills': [],
            'damageDealtToChampions': [],
            'damageTaken': [],
            'goldEarned': [],
            'minionsKilled': [],
        }

        for match_id in match_ids:
            print(i, match_ids.index(match_id),match_id)

            # get the match_data from a match ID
            match_data = self.get_match_data(match_id)

            # get player_data from a match ID
            participants = match_data['metadata']['participants']
            player_index = participants.index(self.puuid)
            player_data = match_data['info']['participants'][player_index]

            # extract infomation from match_data & player_data
            matchId = match_id
            gameStartTime = round(match_data['info']['gameStartTimestamp']/1000)
            date = str(datetime.datetime.fromtimestamp(gameStartTime))[:10]
            time = str(datetime.datetime.fromtimestamp(gameStartTime))[10:]
            gameLength = str(datetime.timedelta(seconds=int(player_data['challenges']['gameLength'])))
            mode = match_data['info']['gameMode']
            gameVersion = match_data['info']['gameVersion']
            team = 'blue' if player_data['teamId'] == 100 else 'red'
            win = 1 if player_data['win'] else 0
            earlySurrender = 1 if player_data["gameEndedInEarlySurrender"] else 0
            surrender = 1 if player_data["gameEndedInSurrender"] else 0
            summonerName = player_data['summonerName']
            summonerLevel = player_data['summonerLevel']
            champion = player_data['championName']
            k = player_data['kills']
            d = player_data['deaths']
            a = player_data['assists']
            if d != 0:
                kda = round((k+a)/d, 2)
            elif d == 0:
                kda = 'N/A'
            quadraKills = player_data['quadraKills']
            pentaKills = player_data['pentaKills']
            totalDamageDealtToChampions = player_data['totalDamageDealtToChampions']
            totalDamageTaken = player_data['totalDamageTaken']
            goldEarned = player_data['goldEarned']
            totalMinionsKilled = player_data['totalMinionsKilled']
            
            # add them to data
            data['matchId'].append(matchId)
            data['gameStartTime'].append(gameStartTime)
            data['date'].append(date)
            data['time'].append(time)
            data['gameLength'].append(gameLength)
            data['mode'].append(mode)
            data['gameVersion'].append(gameVersion)
            data['team'].append(team)
            data['win'].append(win)
            data['earlySurrender'].append(earlySurrender)
            data['surrender'].append(surrender)
            data['summonerName'].append(summonerName)
            data['summonerLevel'].append(summonerLevel)
            data['champion'].append(champion)
            data['kills'].append(k)
            data['deaths'].append(d)
            data['assists'].append(a)
            data['KDA'].append(kda)
            data['quadraKills'].append(quadraKills)
            data['pentaKills'].append(pentaKills)
            data['damageDealtToChampions'].append(totalDamageDealtToChampions)
            data['damageTaken'].append(totalDamageTaken)
            data['goldEarned'].append(goldEarned)
            data['minionsKilled'].append(totalMinionsKilled)
        
        gatheredDf = pd.DataFrame(data)
        if match_ids == []:
            gameStartTime = 0
        return gatheredDf, gameStartTime
    #=============================================================================================================
    #=============================================================================================================
    def master_function(self):
        # reset the self.df to a blank dataframe
        self.df = pd.DataFrame()
        endTime = self.endTime

        # determine the round number
        if self.no_games < 100: # one run only
            i = 100 # counter i
        elif self.no_games == 100: # wants continuous runs until all match found
            i = 1 # counter i
        
        # initialise match_ids != []
        match_ids = [0]

        # Loop: assume max round to run is 100
        while i <= 100:
            # get match_ids (max. 100 matches a time)
            match_ids = self.get_match_ids(endTime)

            # break if no matches return
            if match_ids == []:
                break

            # gather data of each match in the list
            returnDf = self.gather_all_Data(match_ids, i)

            # combine the df this round with last round 
            self.df = pd.concat([self.df, returnDf[0]])
            endTime = returnDf[1] - 1 # endTime keeps decrease until match_ids == [] then break the loop
            i += 1 # complete of one match_ids list)
        print('Match data gathered!') 
    #=============================================================================================================
    #=============================================================================================================
    def statistics(self):
        if self.df.empty == True:
            print('df is empty, unable to run statistics(), please gather data (mode 1) or read old data from excel (mode 2) first.')
        elif self.df.empty == False:
            # Find the averages: numeric_only stops it trying to average the "champion" column
            mean = self.df.mean(numeric_only=True).to_frame()
            mean = mean.rename(columns={0: "mean"})
            mean = mean.rename(index={'win': 'winRate'})
            mean.index.name = '0'
            mean = mean.drop(['gameStartTime','summonerLevel','earlySurrender','surrender'])
            mean = mean.mul([100,1,1,1,1,1,1,1,1,1,1], axis='index')
            mean = mean.round(2)
            new_row = pd.DataFrame({len(self.df.index)},index=['gameCount']).rename(columns={0: "mean"})
            mean = pd.concat([new_row, mean],axis='rows')

            # Get the averages per champion
            dfDrop = self.df
            dfDrop = dfDrop.drop(columns=['gameStartTime','summonerLevel'])
            dfDrop = dfDrop.rename(columns={'win':'winRate'})
            gpByChamMean = dfDrop.groupby('champion').mean(numeric_only=True).round(3)
            championCount = dfDrop.groupby('champion').size()
            gpByChamMean = pd.concat([championCount, gpByChamMean],axis='columns')
            gpByChamMean = gpByChamMean.rename(columns={0: "count"})
            gpByChamMean['winRate'] = gpByChamMean['winRate']*100
            gpByChamMean = gpByChamMean.sort_values(by=['count'], ascending=False)

            self.stat = {'mean' : mean , 'gpByChamMean' : gpByChamMean}
            print('Statistic ready!')
    #=============================================================================================================
    #=============================================================================================================
    def adjustExcelColumnWidth():
        wb = load_workbook('output.xlsx')

        for ws in ['data', 'mean', 'gpByChamMean']:
            for col_number in range(1, wb[ws].max_column + 1):
                col_letter = get_column_letter(col_number)
                max_width = 0
                for row_number in range(1, wb[ws].max_row + 1): # wb[ws].max_row + 1
                    # print(wb[ws][f'{letter}{row_number}'].value)
                    if len(str(wb[ws][f'{col_letter}{row_number}'].value)) > max_width:
                        max_width = len(str(wb[ws][f'{col_letter}{row_number}'].value))
                wb[ws].column_dimensions[col_letter].width = max_width + 4

        wb.save('output.xlsx')
        print("Excel column width adjusted!")
    #=============================================================================================================
    #=============================================================================================================
    # output as excel & csv
    def outputAsExcel(self):
        if self.df.empty == True:
            print('df is empty, unable to run outputAsExcel() , please gather data (mode 1) or read old data from excel (mode 2) first.')
        elif self.stat == {}:
            print('stat is empty, please generate statistic (mode 3) first.')
        elif self.df.empty == False and self.stat != {}:
            fileName = 'output'
            mean = self.stat['mean']
            gpByChamMean = self.stat['gpByChamMean']
            
            with pd.ExcelWriter(fileName + ".xlsx") as writer:
                self.df.set_index("matchId", inplace=True)
                self.df.to_excel(writer, sheet_name="data")
                # auto_adjust_xlsx_column_width(self.df, writer, sheet_name="data", margin=3)

                mean.to_excel(writer, sheet_name="mean")
                # auto_adjust_xlsx_column_width(mean, writer, sheet_name="mean", margin=3)

                gpByChamMean.to_excel(writer, sheet_name="gpByChamMean") 
                # auto_adjust_xlsx_column_width(gpByChamMean, writer, sheet_name="gpByChamMean", margin=3)

            # adjust column width
            Data.adjustExcelColumnWidth()

            print('Output excel succeed!')
    #=============================================================================================================
    #=============================================================================================================
    def updateOldExcel(self):
        oldDf = pd.read_excel('./output.xlsx',sheet_name='data')
        self.startTime = oldDf['gameStartTime'][0] # lastest gameTime
        self.master_function() # gather matches after lastest gameTime
        oldDf = oldDf.drop([0]) # the lastest match of oldDf will repeat, drop that repeated match
        self.df = pd.concat([self.df, oldDf])
        print('Old excel updated!')
    #=============================================================================================================
    #=============================================================================================================
    def timeInputToEpoch(dateInput):
        year = int(str(dateInput)[:4])
        month = int(str(dateInput)[5]) if str(dateInput)[4] == 0 else int(str(dateInput)[4:6])
        day = int(str(dateInput)[7]) if str(dateInput)[6] == 0 else int(str(dateInput)[6:8])

        datetime_obj = datetime.datetime(year,month,day)
        epoch = round(datetime_obj.timestamp())
        return epoch
    #=============================================================================================================
    #=============================================================================================================
    #=============================================================================================================
    #==========================================For mode choosing==================================================
    #=============================================================================================================
    #=============================================================================================================
    #=============================================================================================================
    def chooseMode(self):
        # check API key is vaild
        self.checkApiKey()

        while True:
            self.printMenu()
            mode = input('please input mode:\n')
            match mode:
                case '0': self.mode0()
                case '1': self.mode1()
                case '2': self.mode2()
                case '3': self.mode3()
                case '4': self.mode4()
                case _  : print('unrecognized input, please try again!')
    #=============================================================================================================
    #=============================================================================================================
    def printMenu(self):
        menu = """==============================
mode   function
0      terminate program
1      gather match data
2      output as excel
3      update existing excel
4      change variables
=============================="""
        print(menu)
    #=============================================================================================================
    #=============================================================================================================
    def printVariableMenu(self):
        print(f'''============================================================
       Current Variable
0      Return to Menu
1      summoner_name = {self.summoner_name}
2      puuid = {self.puuid} (don't change this, change by summoner_name)
3      region = {self.region}
4      mass_region = {self.mass_region} 
5      no_games = {self.no_games}
6      queue_id = {self.queue_id}
7      startTime = {datetime.datetime.fromtimestamp(self.startTime)}
8      endTime = {datetime.datetime.fromtimestamp(self.endTime)}
9      api_key = {self.api_key}
============================================================''')
    #=============================================================================================================
    #=============================================================================================================
    def changeVariable(self):
        num = '100' # 100 means looping
        while num != '0':
            self.printVariableMenu()
            num = input('please input the number (0-7) of variable to change:\n')
            match num:
                case '0': 
                    num = '0'
                    print('0 received, return to menu.')
                case '1': 
                    newVar = input(f'''please enter new summoner_name or choose from 1-7 (0 to exit):
    1: 大涼粉       2: 國士無雙十三面聽    3:笨蛋變態吵死了
    4: 全Loss 35P   5: OnewaySubaru486   6: 滑她咩墮
    7: 下北澤奇跡大天使\n''')
                    if newVar != '0':
                        match newVar:
                            case '1':newVar = '大涼粉'
                            case '2':newVar = '國士無雙十三面聽'
                            case '3':newVar = '笨蛋變態吵死了'
                            case '4':newVar = '全Loss 35P'
                            case '5':newVar = 'OnewaySubaru486'
                            case '6':newVar = '滑她咩墮'
                            case '7':newVar = '下北澤奇跡大天使'
                        self.summoner_name = newVar
                        self.get_puuid()
                case '2':
                    newVar = input('please enter new puuid (0 to exit):\n')
                    if newVar != '0':
                        self.puuid = newVar
                        self.updateSummonerName()
                case '3': 
                    newVar = input('please enter new region (0 to exit):\n')
                    if newVar != '0':
                        self.region = newVar
                case '4':
                    newVar = input('please enter new mass_region (0 to exit):\n')
                    if newVar != '0':
                        self.mass_region = newVar
                case '5': 
                    newVar = input('please enter new no_games (1-100, 100 means all matches in the time range)(0 to exit):\n')
                    if newVar != '0':
                        self.no_games = int(newVar)
                case '6': 
                    newVar = input('please enter new queue_id (only support 450 ARAM)(0 to exit):\n')
                    if newVar != '0':
                        self.queue_id = int(newVar)
                case '7': 
                    newVar = input('please enter new startTime (YYYYMMDD)(0 to exit):\n')
                    if newVar != '0':
                        self.startTime = self.timeInputToEpoch(newVar)
                case '8': 
                    newVar = input('please enter new endTime (YYYYMMDD)(0 to exit):\n')
                    if newVar != '0':
                        self.endTime = self.timeInputToEpoch(newVar)
                case '9': 
                    newVar = input('please enter new api_key(0 to exit):\n')
                    if newVar != '0':
                        self.api_key = self.checkApiKey(newVar)
                case _: print('unrecognized input, please try again!')
    #=============================================================================================================
    #=============================================================================================================
    def mode0(self):
        print('0 received, program end.')
        exit()
    #=============================================================================================================
    #=============================================================================================================
    def mode1(self):
        print('1 received')
        # gather match data
        self.master_function()
    #=============================================================================================================
    #=============================================================================================================
    def mode2(self):
        print('2 received')
        # print statistics: analyse the data
        self.statistics()
        # output as excel
        self.outputAsExcel()
    #=============================================================================================================
    #=============================================================================================================
    def mode3(self):
        print('3 received')
        # Update the lastest match to old excel
        if os.path.exists('./output.xlsx') == False:
            print('\'output.xlsx\' not found')
        elif os.path.exists('./output.xlsx'):
            self.updateOldExcel()
            self.statistics()
            self.outputAsExcel()
    #=============================================================================================================
    #=============================================================================================================
    def mode4(self):
        print('4 received')
        self.changeVariable()
    #=============================================================================================================
    #=============================================================================================================
    # NOTE
    # Tuning format of df by pandas: .astype: convert "win" from boolean to integer & control decimal by .round
    # df['win'] = df['win'].astype(int)